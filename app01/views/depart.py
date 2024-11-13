from django.shortcuts import render, redirect
from app01 import models
from django.db.models import Q
from app01.utils.pagination import Pagination
from app01.utils.form import UserModelForm, PrettyModelForm, PrettyEditModelForm

def depart_list(request):
    """
    部门列表页面，支持搜索功能
    """
    # 获取用户输入的查询关键字
    query = request.GET.get('q', '')
    # 获取所有部门数据，如果有查询关键字则进行过滤
    queryset = models.Department.objects.all()
    if query:
        queryset = queryset.filter(Q(title__icontains=query))

    # 使用分页工具
    page_object = Pagination(request, queryset, page_size=20)
    context = {
        "queryset": page_object.page_queryset,  # 分页后的数据
        "page_string": page_object.html(),  # 页码显示
        "query": query  # 搜索关键词
    }
    return render(request, 'depart_list.html', context)


def depart_add(request):
    """
    添加新部门
    """
    if request.method == "GET":
        return render(request, 'depart_add.html')

    # 获取用户通过POST提交的部门名称
    title = request.POST.get("title")
    # 保存新部门到数据库
    models.Department.objects.create(title=title)
    # 重定向到部门列表页面
    return redirect("/depart/list/")


def depart_delete(request):
    """
    删除部门
    """
    # 从URL中获取要删除的部门ID
    nid = request.GET.get('nid')
    # 根据ID删除对应的部门
    models.Department.objects.filter(id=nid).delete()
    # 重定向回部门列表页面
    return redirect("/depart/list/")


def depart_edit(request, nid):
    """
    修改部门信息
    """
    if request.method == "GET":
        # 根据ID获取部门数据
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html', {"row_object": row_object})

    # 获取用户修改后的部门名称
    title = request.POST.get("title")
    # 更新数据库中的部门信息
    models.Department.objects.filter(id=nid).update(title=title)
    # 重定向到部门列表页面
    return redirect("/depart/list/")


def depart_multi(request):
    """
    批量导入部门（通过Excel文件上传）
    """
    from openpyxl import load_workbook

    # 获取用户上传的文件
    file_object = request.FILES.get("exc")
    # 使用openpyxl加载Excel文件
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 循环读取每一行数据，从第二行开始
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        # 检查部门是否已存在，避免重复导入
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            # 不存在时创建新的部门
            models.Department.objects.create(title=text)

    # 重定向到部门列表页面
    return redirect('/depart/list/')